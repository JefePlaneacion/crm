import pandas as pd
import requests
from datetime import datetime
import numpy as np
import json, ast
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator
from pathlib import Path




url = 'https://tg.toscanagroup.com.co/api_powerbi.php'
today_dateo = datetime.now().date()
params = {
     "auth": {
        "user": "jorge.contreras",
        "pass": "EstebanGrey1704*"
    },
    "data": {
        "type":"APROBADOS",
        "f_inicio": "2024-05-01",
        "f_fin": today_dateo.strftime("%Y-%m-%d")
    }
}

# Realizar la solicitud POST a la API
response = requests.post(url=url, json=params)
today_date = datetime.now().date()

if response.status_code == 200:
    data=response.json()
    df = pd.DataFrame(data)

df_final=df.copy()['registros_aprobados'].apply(pd.Series)

df_final.to_excel("pedidosInicial.xlsx", index=False)



# ============================================
# 2. PROCESAMIENTO DE PRODUCTOS
# ============================================
df_producto = df_final[['pedido', 'cliente','tipo_producto', 'oportunidad', 'razon_anulacion', 'f_cotizacion', 'f_pedido', 'f_comercial', 'f_diseno', 'f_produccion', 'f_despachos', 'f_instalacion',
                        'fe_comercial', 'fe_diseno', 'fe_produccion', 'fe_despachos', 'fe_instalacion','estado', 'regional', 'productos']].copy()

df_producto.to_excel("pedidosProductosInicial.xlsx", index=False)

def parse_product_cell(x):
    """
    Función para parsear la celda de productos que puede venir en diferentes formatos
    """
    # 1) Ya es lista
    if isinstance(x, list):
        return x
    
    # 2) NaN / None
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return []
    
    s = str(x).strip()

    # 3) Intenta JSON directo
    try:
        return json.loads(s)
    except Exception:
        pass

    # 4) Intenta literal de Python (acepta comillas simples)
    try:
        return ast.literal_eval(s)
    except Exception:
        pass

    # 5) Limpieza mínima y reintentos
    s2 = (s.replace("None", "null")
           .replace("True", "true")
           .replace("False", "false")
           .replace("'", "'")
           .replace(""", '"').replace(""", '"'))
    
    for candidate in (s2, s2.replace("'", '"')):
        try:
            return json.loads(candidate)
        except Exception:
            continue

    # 6) Si nada funciona: devuelve lista vacía para no romper
    return []


# Parsear la columna de productos
df_producto.loc[:, "productos"] = df_producto["productos"].apply(parse_product_cell)

# Explode y normalizar
df_expandido = df_producto.explode("productos", ignore_index=True)
df_expandido.loc[df_expandido["productos"].isna(), "productos"] = {}

# Concatenar con las columnas expandidas de productos
df_final_prod = pd.concat(
    [df_expandido.drop(columns=["productos"]),
     df_expandido["productos"].apply(lambda d: d if isinstance(d, dict) else {}).apply(pd.Series)],
    axis=1
)

df_final_prod["valor_total"] = pd.to_numeric(df_final_prod["valor_total"], errors="coerce")


#BASE DE DATOS PARA PLAN DE PRODUCCIÓN

df_final_plan_prod = df_final_prod.copy()  #BASE DE DATOS PARA PLAN DE PRODUCCIÓN

df_final_plan_prod=df_final_plan_prod[df_final_plan_prod['tipo_producto'] !='NO CONFORME']

df_final_plan_prod=df_final_plan_prod[df_final_plan_prod['id_producto'] !='Id_Caratula']

s = df_final_plan_prod['producto'].str.split(' - ', n=1).str.get(1)
df_final_plan_prod['producto'] = s.fillna(df_final_plan_prod['producto'])

df_final_plan_prod['producto'] = df_final_plan_prod['producto'].str.strip()

df_final_plan_prod=df_final_plan_prod[df_final_plan_prod['oportunidad'] !=-1]


conditions = [
    df_final_plan_prod["estado"]=="Aprobado Comercial",
    df_final_plan_prod["estado"]=="En Proceso Comercial",
    df_final_plan_prod["estado"]=="Detenido Comercial",
    df_final_plan_prod["estado"]=="Aprobado Diseño",
    df_final_plan_prod["estado"]=="En Proceso Diseño",
    df_final_plan_prod["estado"]=="Detenido Diseño",
    df_final_plan_prod["estado"]=="En Proceso Producción",
    df_final_plan_prod["estado"]=="Detenido Producción",
    df_final_plan_prod["estado"]=="Aprobado Pedido",
    
]
choices = ["DISEÑO", "COMERCIAL", "COMERCIAL", "PRODUCCION", "DISEÑO","DISEÑO","PRODUCCION","PRODUCCION","COMERCIAL"]

df_final_plan_prod["estado_crm"] = np.select(conditions, choices, default="FINALIZADO") 



def actualizar_base_datos_seguro(
    ruta_excel: str,
    df: pd.DataFrame,
    hoja_objetivo: str = "baseDatos",
    fila_header: int = 1,
    columnas_datos_whitelist: list[str] | None = None,  # columnas que SÍ puedo escribir
    detectar_formulas_hasta_filas: int = 5,             # cuántas filas de datos escanear para detectar fórmulas
    limpiar_solo_columnas_datos: bool = True,           # NO limpiar columnas que no estén en whitelist
    rellenar_formulas: bool = False,                    # por defecto NO tocamos fórmulas
):
    ruta = Path(ruta_excel)
    keep_vba = ruta.suffix.lower() == ".xlsm"

    wb = load_workbook(ruta_excel, keep_vba=keep_vba, data_only=False)  # data_only=False para leer fórmulas
    if hoja_objetivo not in wb.sheetnames:
        raise ValueError(f"No existe la hoja '{hoja_objetivo}' en '{ruta_excel}'.")
    ws = wb[hoja_objetivo]

    # 1) Leer encabezados
    headers = {}
    col = 1
    # límite prudente
    while col <= 500:
        v = ws.cell(row=fila_header, column=col).value
        if v is not None and str(v).strip():
            headers[str(v).strip()] = col
        col += 1

    # 2) Definir columnas que escribiré (whitelist)
    if columnas_datos_whitelist is None:
        columnas_datos = [c for c in df.columns if c in headers]  # solo las que existan en la hoja
    else:
        columnas_datos = [c for c in columnas_datos_whitelist if c in headers]

    # 3) Detectar columnas con fórmula (excluirlas siempre)
    fila_datos_ini = fila_header + 1
    columnas_con_formula = set()
    if detectar_formulas_hasta_filas and ws.max_row >= fila_datos_ini:
        to_row = min(ws.max_row, fila_datos_ini + max(detectar_formulas_hasta_filas - 1, 0))
        for nombre_col, cidx in headers.items():
            for r in range(fila_datos_ini, to_row + 1):
                val = ws.cell(row=r, column=cidx).value
                if isinstance(val, str) and val.startswith("="):
                    columnas_con_formula.add(nombre_col)
                    break

    # Excluir de la whitelist cualquier columna con fórmula detectada
    columnas_datos = [c for c in columnas_datos if c not in columnas_con_formula]

    # 4) Validación: que el DF tenga esas columnas
    faltantes_en_df = [c for c in columnas_datos if c not in df.columns]
    if faltantes_en_df:
        raise ValueError(f"Estas columnas esperadas no están en df: {faltantes_en_df}")

    # 5) Limpieza suave SOLO en columnas_datos (nunca en columnas con fórmula ni en otras hojas)
    if limpiar_solo_columnas_datos and ws.max_row >= fila_datos_ini:
        for nombre_col in columnas_datos:
            cidx = headers[nombre_col]
            for r in range(fila_datos_ini, ws.max_row + 1):
                ws.cell(row=r, column=cidx).value = None

    # 6) Escribir df en columnas_datos
    for i, (_, row) in enumerate(df.iterrows(), start=fila_datos_ini):
        for nombre_col in columnas_datos:
            cidx = headers[nombre_col]
            ws.cell(row=i, column=cidx).value = row[nombre_col]

    # 7) (Opcional) Rellenar fórmulas SOLO en columnas que ya tenían fórmula
    if rellenar_formulas and columnas_con_formula:
        nfilas_nuevas = len(df)
        last_row = fila_datos_ini + max(nfilas_nuevas - 1, 0)
        for nombre_col in (set(headers) & columnas_con_formula):
            cidx = headers[nombre_col]
            # buscar una celda "plantilla" con fórmula en el bloque superior
            plantilla = None
            plantilla_coord = None
            for r in range(fila_datos_ini, fila_datos_ini + detectar_formulas_hasta_filas):
                val = ws.cell(row=r, column=cidx).value
                if isinstance(val, str) and val.startswith("="):
                    plantilla = val
                    plantilla_coord = ws.cell(row=r, column=cidx).coordinate
                    break
            if not plantilla:
                continue  # no se propaga nada si no existe plantilla

            for r in range(fila_datos_ini, last_row + 1):
                destino = ws.cell(row=r, column=cidx)
                nueva = Translator(plantilla, origin=plantilla_coord).translate_formula(destino.coordinate)
                destino.value = nueva

    # 8) Guardar (no se tocan otras hojas)
    wb.save(ruta_excel)

# Tras construir df_final_plan_prod ...
ruta = r"C:\Users\JORGE CONTRERAS\proyecto_plan_produccion\pedidosProductosFinal.xlsx"

# Define explícitamente las columnas de DATOS que sí quieres escribir.
# (Las de fórmula NUNCA van aquí.)
columnas_datos = [
    "pedido","cliente","tipo_producto","oportunidad","razon_anulacion","f_cotizacion","f_pedido","f_comercial","f_diseno","f_produccion","f_despachos","f_instalacion","fe_comercial",
    "fe_diseno","fe_produccion","fe_despachos","fe_instalacion","estado","regional","id_producto", "producto", "color","color_estructura","cantidad","largo","alto","proyeccion" ,"valor_total","estado_crm",
    # agrega aquí SOLO columnas que NO tienen fórmulas en baseDatos
]

actualizar_base_datos_seguro(
    ruta_excel=ruta,
    df=df_final_plan_prod,
    hoja_objetivo="baseDatos",
    fila_header=1,
    columnas_datos_whitelist=columnas_datos,   # whitelist estricta
    detectar_formulas_hasta_filas=10,          # detecta fórmulas en primeras 10 filas de datos
    limpiar_solo_columnas_datos=True,          # solo limpia esas columnas
    rellenar_formulas=False                    # pon True SOLO si necesitas auto-fill de fórmulas
)






# -----------------------
# 1) Preparación
# -----------------------
df_final_prod_ped = df_final[['pedido','cliente','tipo_producto','oportunidad','razon_anulacion','estado','regional',
                              'f_cotizacion','f_pedido','f_comercial','f_diseno','f_produccion',
                              'f_despachos','f_instalacion',
                              'fe_comercial','fe_diseno','fe_produccion','fe_despachos','fe_instalacion']].copy()

cols_fecha = [
    'f_cotizacion','f_pedido','f_comercial','f_diseno','f_produccion',
    'f_despachos','f_instalacion','fe_comercial','fe_diseno',
    'fe_produccion','fe_despachos','fe_instalacion'
]

def is_blank(x):
    if pd.isna(x):
        return True
    s = str(x).strip().lower()
    return s == "" or s in {"none", "nan"}

# -----------------------
# 2) Convertir a datetime seguro y calcular estado_crm ANTES de rellenar
# -----------------------
def to_datetime_safe(s: pd.Series) -> pd.Series:
    dt = pd.to_datetime(s, format="%Y-%m-%d %H:%M:%S", errors="coerce")
    mask = dt.isna()
    if mask.any():
        dt.loc[mask] = pd.to_datetime(s[mask], format="%Y-%m-%d", errors="coerce")
    return dt

df_final_prod_ped[cols_fecha] = df_final_prod_ped[cols_fecha].apply(to_datetime_safe)

# Calcular estado_crm ANTES de rellenar fechas (basado en el estado actual)
conditions = [
    df_final_prod_ped["estado"]=="Aprobado Comercial",
    df_final_prod_ped["estado"]=="En Proceso Comercial",
    df_final_prod_ped["estado"]=="Detenido Comercial",
    df_final_prod_ped["estado"]=="Aprobado Diseño",
    df_final_prod_ped["estado"]=="En Proceso Diseño",
    df_final_prod_ped["estado"]=="Detenido Diseño",
    df_final_prod_ped["estado"]=="Aprobado Producción",
    df_final_prod_ped["estado"]=="Aprobado Produccion",
    df_final_prod_ped["estado"]=="En Proceso Producción",
    df_final_prod_ped["estado"]=="Detenido Producción",
    df_final_prod_ped["estado"]=="Aprobado Instalación",
    df_final_prod_ped["estado"]=="En Proceso Instalación",
    df_final_prod_ped["estado"]=="Detenido Instalación",
    df_final_prod_ped["estado"]=="Aprobado Muebles",
    df_final_prod_ped["estado"]=="Aprobado Despacho Muebles",
    df_final_prod_ped["estado"]=="Detenido Despacho Muebles",
    df_final_prod_ped["estado"]=="Aprobado Despachos",
    df_final_prod_ped["estado"]=="En Proceso Despachos",
    df_final_prod_ped["estado"]=="Detenido Despachos",
    df_final_prod_ped["estado"]=="Aprobado Pedido",
    df_final_prod_ped["pedido"]==106671,
    df_final_prod_ped["pedido"]==107242,
    df_final_prod_ped["pedido"]==107243,
    df_final_prod_ped["pedido"]==107634,    
    df_final_prod_ped["pedido"]==108330,
    
]
choices = ["DISEÑO", "COMERCIAL", "COMERCIAL", "PRODUCCION", "DISEÑO","DISEÑO","DESPACHOS","DESPACHOS","PRODUCCION","PRODUCCION","FINALIZADO","INSTALACION","INSTALACION","PRODUCCION","INSTALACION","DESPACHOS","INSTALACION","DESPACHOS","DESPACHOS","COMERCIAL","PRODUCCION","DISEÑO","DISEÑO","COMERCIAL","INSTALACION"]

df_final_prod_ped["estado_crm"] = np.select(conditions, choices, default="FINALIZADO")

# -----------------------
# 3) Rellenar fechas según estado
# -----------------------


maskNPD = (df_final_prod_ped["estado"].astype(str).str.strip().str.upper() == "APROBADO COMERCIAL") & df_final_prod_ped["f_produccion"].isna() & df_final_prod_ped["tipo_producto"]=="MUEBLES"
df_final_prod_ped.loc[maskNPD, 'f_produccion'] = pd.to_datetime(today_date)


maskNPDL = (df_final_prod_ped["estado"].astype(str).str.strip().str.upper() == "APROBADO COMERCIAL") & df_final_prod_ped["f_produccion"].isna() & df_final_prod_ped["tipo_producto"]=="CUBRIMIENTOS"
df_final_prod_ped.loc[maskNPDL, 'f_diseno'] = pd.to_datetime(today_date)


maskDTCC= (
    df_final_prod_ped["estado"].astype(str).str.strip().str.upper()=="EN PROCESO COMERCIAL"
) & df_final_prod_ped["f_comercial"].isna()
df_final_prod_ped.loc[maskDTCC,'f_comercial'] = pd.to_datetime(today_date)



maskDTC = (
    df_final_prod_ped["estado"].astype(str).str.strip().str.upper()=="DETENIDO COMERCIAL"
) & df_final_prod_ped["f_comercial"].isna()
df_final_prod_ped.loc[maskDTC,'f_comercial'] = pd.to_datetime(today_date)



maskDTC = (
    df_final_prod_ped["estado"].astype(str).str.strip().str.upper()=="APROBADO DISEÑO"
)
df_final_prod_ped.loc[maskDTC,'f_produccion'] = pd.to_datetime(today_date)

maskDTCP = (
    df_final_prod_ped["estado"].astype(str).str.strip().str.upper()=="EN PROCESO DISEÑO"
)
df_final_prod_ped.loc[maskDTCP,'f_diseno'] = pd.to_datetime(today_date)

maskDTDS = (
    df_final_prod_ped["estado"].astype(str).str.strip().str.upper()=="DETENIDO DISEÑO"
) & df_final_prod_ped["f_diseno"].isna()
df_final_prod_ped.loc[maskDTDS,'f_diseno'] = pd.to_datetime(today_date)



maskDTCPP = (
    df_final_prod_ped["estado"].astype(str).str.strip().str.upper()=="APROBADO PRODUCCIÓN"
)
df_final_prod_ped.loc[maskDTCPP,'f_despachos'] = pd.to_datetime(today_date)


maskDTPR = (
    df_final_prod_ped["estado"].astype(str).str.strip().str.upper()=="DETENIDO PRODUCCIÓN"
) & df_final_prod_ped["f_produccion"].isna()
df_final_prod_ped.loc[maskDTPR,'f_produccion'] = pd.to_datetime(today_date)

maskDTPRO = (
    df_final_prod_ped["estado"].astype(str).str.strip().str.upper()=="DETENIDO PRODUCCION"
) & df_final_prod_ped["f_produccion"].isna()
df_final_prod_ped.loc[maskDTPRO,'f_produccion'] = pd.to_datetime(today_date)


maskDTCPEP = (
    df_final_prod_ped["estado"].astype(str).str.strip().str.upper()=="EN PROCESO PRODUCCION"
)
df_final_prod_ped.loc[maskDTCPEP,'f_produccion'] = pd.to_datetime(today_date)

maskDTCPEPT = (
    df_final_prod_ped["estado"].astype(str).str.strip().str.upper()=="APROBADO DESPACHOS"
)
df_final_prod_ped.loc[maskDTCPEPT,'f_instalacion'] = pd.to_datetime(today_date)

maskDTCPEPTM = (
    df_final_prod_ped["estado"].astype(str).str.strip().str.upper()=="APROBADO DESPACHO MUEBLES"
)
df_final_prod_ped.loc[maskDTCPEPTM,'f_instalacion'] = pd.to_datetime(today_date)


maskDTDP = (
    df_final_prod_ped["estado"].astype(str).str.strip().str.upper()=="DETENIDO DESPACHOS"
)
df_final_prod_ped.loc[maskDTDP,'f_despachos'] = pd.to_datetime(today_date)
maskDTDPM = (
    df_final_prod_ped["estado"].astype(str).str.strip().str.upper()=="DETENIDO DESPACHO MUEBLES"
)
df_final_prod_ped.loc[maskDTDPM,'f_despachos'] = pd.to_datetime(today_date)
maskDTDPMPD = (
    df_final_prod_ped["estado"].astype(str).str.strip().str.upper()=="EN PROCESO DESPACHOS"
)
df_final_prod_ped.loc[maskDTDPMPD,'f_despachos'] = pd.to_datetime(today_date)
maskDTDPMPDM = (
    df_final_prod_ped["estado"].astype(str).str.strip().str.upper()=="EN PROCESO DESPACHO MUEBLES"
)
df_final_prod_ped.loc[maskDTDPMPDM,'f_despachos'] = pd.to_datetime(today_date)

maskDTDPMPDMI = (
    df_final_prod_ped["estado"].astype(str).str.strip().str.upper()=="EN PROCESO INSTALACIÓN"
)
df_final_prod_ped.loc[maskDTDPMPDMI,'f_instalacion'] = pd.to_datetime(today_date)

maskDTDPMPDMICT = (
    df_final_prod_ped["estado"].astype(str).str.strip().str.upper()=="COTIZADO"
) & (df_final_prod_ped["pedido"]==106671)
df_final_prod_ped.loc[maskDTDPMPDMICT,'f_produccion'] = pd.to_datetime(today_date)
maskDTDPMPDMICTO = (
    df_final_prod_ped["estado"].astype(str).str.strip().str.upper()=="COTIZADO"
) & (df_final_prod_ped["pedido"]==107242)
df_final_prod_ped.loc[maskDTDPMPDMICTO,'f_diseno'] = pd.to_datetime(today_date)
df_final_prod_ped.loc[maskDTDPMPDMICT,'f_produccion'] = pd.to_datetime(today_date)
maskDTDPMPDMICTOP = (
    df_final_prod_ped["estado"].astype(str).str.strip().str.upper()=="COTIZADO"
) & (df_final_prod_ped["pedido"]==107243)
df_final_prod_ped.loc[maskDTDPMPDMICTOP,'f_diseno'] = pd.to_datetime(today_date)
maskDTDPMPDMICTOPT = (
    df_final_prod_ped["estado"].astype(str).str.strip().str.upper()=="COTIZADO"
) & (df_final_prod_ped["pedido"]==107634)
df_final_prod_ped.loc[maskDTDPMPDMICTOPT,'f_comercial'] = pd.to_datetime(today_date)
maskDTDPMPDMICTOPTL = (
    df_final_prod_ped["estado"].astype(str).str.strip().str.upper()=="COTIZADO"
) & (df_final_prod_ped["pedido"]==108330)
df_final_prod_ped.loc[maskDTDPMPDMICTOPTL,'f_instalacion'] = pd.to_datetime(today_date)


maskDTIN = (
    df_final_prod_ped["estado"].astype(str).str.strip().str.upper()=="DETENIDO INSTALACIÓN"
) & df_final_prod_ped["f_instalacion"].isna()
df_final_prod_ped.loc[maskDTIN,'f_instalacion'] = pd.to_datetime(today_date)


# -----------------------
# 4) Formato D/M/Y
# -----------------------
def formato_DMY(s:pd.Series) -> pd.Series:
    out = (
        s.dt.day.astype("Int64").astype(str) + "/" +
        s.dt.month.astype("Int64").astype(str) + "/" +
        s.dt.year.astype("Int64").astype(str)
    )
    return out.where(s.notna(), None)

df_final_prod_ped[cols_fecha] = df_final_prod_ped[cols_fecha].apply(formato_DMY)

df_final_prod_ped["cliente"] = df_final_prod_ped["cliente"].fillna("NO CONFORME")

# -----------------------
# 5) Exportar
# -----------------------
df_final_prod_ped.to_excel("pedidosFinal.xlsx", index=False)
